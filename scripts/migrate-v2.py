#!/usr/bin/env python3
"""Migration old Verknuepfungen schema -> new long-format (v2), pilot scope.

Reads the legacy multi-sheet ``M3GIM-Verknuepfungen.xlsx`` (columns
signatur/Folio/datenpunkt_id/typ/name/rolle/anmerkung) and emits the new
long-format draft for one signature prefix (default UAKUG/NIM_137).

Design constraints agreed with the operator:
- ``aktivitaet_id`` stays EMPTY. Activity grouping is left to the human pass;
  the migration never guesses which performance a row belongs to.
- Every output row is a pure, traceable transformation of one source row.
  No cross-row merging, no deduplication, no invented content.
- Cell cleaning only: split composite cells, split money, split
  funktion/rolle, normalize dates where ISO-parseable, map controlled
  vocab via explicit alias tables. Anything not deterministically mappable
  is carried verbatim and flagged.

Tooling columns (stripped before Google-Sheets import):
- ``_block`` : within-folio cohesion key. Rows produced from ONE source row
  share a _block, so the human can assign one aktivitaet_id per participation
  without re-reading the source. Pure source-row provenance, not inference.
- ``_prov``  : source coordinate ``sheet:rowindex`` for round-trip testing.
- ``_flag``  : review hint (erwaehnung, aktivitaet-trigger, name-form,
  datum-format, vokab, komposit).

Run:
    python scripts/migrate-v2.py [SIGNATUR-PREFIX]
Outputs to data/migration/ : <prefix>.xlsx, <prefix>.csv, migration-report.md
Self-tests run at the end and fail loud on any unexplained data loss.
"""
import csv
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "google-spreadsheet" / "M3GIM-Verknüpfungen.xlsx"
OUTDIR = ROOT / "data" / "migration"

OUT_COLS = ["archivsignatur", "Folio", "aktivitaet_id", "typ", "value",
            "anmerkung", "_block", "_prov", "_flag"]

# --- controlled vocab: old rolle -> funktion (canonical, lowercased) --------
# Aliases fix known legacy typos; aliased rows are flagged so the fix is visible.
FUNKTION_ALIASES = {
    "maskenbidner:in": "maskenbildner:in",
}
FUNKTION_ROLES = {
    "komponist:in", "sänger:in", "dirigent:in", "regisseur:in",
    "choreograph:in", "ausstatter:in", "kostümbildner:in", "maskenbildner:in",
    "maskenbidner:in", "technische leitung", "beleuchter:in", "chorleiter:in",
    "bühnenleiter:in", "repetitor:in", "verfasser:in", "herausgeber:in",
    "interpret:in", "leitung", "agent:in", "vermittler:in", "adressat:in",
    "empfänger:in", "absender:in", "unterzeichner:in", "veranstalter:in",
    "vertragspartner:in", "chor", "fluggesellschaft",
}
# rolle values that are NOT a function but a relationship qualifier
MENTION_ROLES = {"erwähnt", "erwaehnt"}
# rolle values that mark an activity trigger (anchor a performance/event)
TRIGGER_ROLES = {"aufführung", "auffuehrung", "auftritt", "gastspiel",
                 "premiere", "probe", "aufnahme"}
# rolle that turns an institution row into an event row
EVENT_ROLES = {"rahmenveranstaltung"}
# ort sub-roles that are not the performance location (mobility / context)
ORT_OTHER_ROLES = {"zielort", "abreiseort", "absendeort", "entstehungsort",
                   "vertragsort", "erwähnt", "erwaehnt"}

ISO_DATE = re.compile(r"^\d{4}(-\d{2}(-\d{2})?)?$")


def norm(v):
    return "" if v is None else str(v).strip()


def lower(v):
    return norm(v).lower()


def name_tokens(raw):
    """Token set for name matching: lowercased, dots->space, comma stripped."""
    return frozenset(
        t for t in re.split(r"[\s,]+", raw.lower().replace(".", " ")) if t
    )


def has_replacement_char(s):
    return "�" in s


# --------------------------------------------------------------------------- #
# loading
# --------------------------------------------------------------------------- #
def load_source(prefix):
    """Return list of source rows as dicts with provenance, filtered to prefix."""
    wb = openpyxl.load_workbook(SRC, read_only=True)
    rows = []
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        hdr = [lower(h) for h in next(it)]
        for i, r in enumerate(it):
            sig = norm(r[0])
            if not sig or not sig.startswith(prefix):
                continue
            d = dict(zip(hdr, r))
            # skip fully empty spreadsheet artifact rows (no data at all)
            if not any(norm(d.get(k)) for k in ("typ", "name", "rolle", "anmerkung")):
                continue
            rows.append({
                "sig": sig,
                "folio": norm(d.get("folio")),
                "typ": norm(d.get("typ")),
                "name": norm(d.get("name")),
                "rolle": norm(d.get("rolle")),
                "anmerkung": norm(d.get("anmerkung")),
                "sheet": ws.title,
                "prov": f"{ws.title}:{i}",
            })
    wb.close()
    return rows


def folio_person_index(rows):
    """canonical person names per folio (from typ=person rows), for name match."""
    idx = defaultdict(dict)  # folio -> {token_set: canonical_name}
    for r in rows:
        if lower(r["typ"]) == "person" and r["name"]:
            idx[r["folio"]][name_tokens(r["name"])] = r["name"]
    return idx


def canonical_person(folio, raw, pidx):
    """Resolve 'Vorname Nachname' to the folio's 'Nachname, Vorname' form."""
    toks = name_tokens(raw)
    cand = pidx.get(folio, {})
    if toks in cand:
        return cand[toks], False  # matched, not flagged
    return raw, True  # keep verbatim, flag name-form


# --------------------------------------------------------------------------- #
# per-row transformation
# --------------------------------------------------------------------------- #
def emit(out, src, block, typ, value, anm="", flag=""):
    out.append({
        "archivsignatur": src["sig"],
        "Folio": src["folio"],
        "aktivitaet_id": "",
        "typ": typ,
        "value": value,
        "anmerkung": anm,
        "_block": block,
        "_prov": src["prov"],
        "_flag": flag,
        "_sheet": src["sheet"],
    })


def transform_row(src, block, pidx, out):
    typ = lower(src["typ"])
    rolle = lower(src["rolle"])
    name = src["name"]
    anm = src["anmerkung"]
    flags = []

    # ---- composite: "datum, werk" ---------------------------------------
    if typ == "datum, werk":
        head, sep, tail = name.partition(",")
        dval, dflag = norm_date(head.strip())
        emit(out, src, block, "datum", dval, anm, ("komposit " + dflag).strip())
        if tail.strip():
            emit(out, src, block, "werk", tail.strip(), anm, "komposit")
        return

    # ---- money composite: "einnahmen|ausgaben|summe, waehrung" -----------
    if typ in ("einnahmen, währung", "ausgaben, währung", "summe, währung"):
        base = typ.split(",")[0].strip()
        amount, curr = split_money(name)
        canon = FUNKTION_ALIASES.get(rolle, rolle)
        # payment qualifier (abendgage, gage) is not a person function -> note it
        qual = "" if not rolle or canon in FUNKTION_ROLES else rolle
        emit(out, src, block, base, amount, join_anm(anm, qual), "geld komposit")
        if curr:
            emit(out, src, block, "währung", curr, "", "komposit")
        if canon in FUNKTION_ROLES:
            emit_funktion(out, src, block, rolle)
        return

    # ---- composite: "ort, datum" ----------------------------------------
    if typ == "ort, datum":
        parts = [p.strip() for p in name.split(",")]
        place = parts[0] if parts else name
        date = parts[1] if len(parts) > 1 else ""
        ptyp = "auffuehrungsort" if rolle in TRIGGER_ROLES or rolle == "auffuehrungsort" else "ort"
        emit(out, src, block, ptyp, place, anm, "komposit")
        if date:
            dval, dflag = norm_date(date)
            emit(out, src, block, "datum", dval, anm, ("komposit " + dflag).strip())
        return

    # ---- composite: "rolle, Vorname Nachname Saenger*in" ----------------
    if typ.startswith("rolle,"):
        head, sep, tail = name.partition(",")
        partie = head.strip()
        person_raw = tail.strip()
        emit(out, src, block, "rolle", partie, anm, "komposit")
        if person_raw:
            canon, nameflag = canonical_person(src["folio"], person_raw, pidx)
            f = "komposit" + (" name-form" if nameflag else "")
            emit(out, src, block, "person", canon, anm, f.strip())
            # the typ label declares "Saenger*in"; take an explicit function
            # role from the rolle column if present, else the declared singer role
            fn = FUNKTION_ALIASES.get(rolle, rolle)
            emit(out, src, block, "funktion",
                 fn if fn in FUNKTION_ROLES else "sänger:in", "", "komposit")
        return

    # ---- bare partie row (typ=rolle) ------------------------------------
    if typ == "rolle":
        flag = "erwaehnung" if rolle in MENTION_ROLES else (
            "aktivitaet-trigger" if rolle in TRIGGER_ROLES else "")
        canon = FUNKTION_ALIASES.get(rolle, rolle)
        # a function role on a bare partie row pertains to the (unnamed) singer
        qual = "" if rolle in MENTION_ROLES | TRIGGER_ROLES or canon in FUNKTION_ROLES else rolle
        emit(out, src, block, "rolle", name, join_anm(anm, qual), flag)
        emit_funktion(out, src, block, rolle)
        return

    # ---- date row (typ=Datum) -------------------------------------------
    if typ == "datum":
        dval, dflag = norm_date(name)
        # the legacy rolle (absendedatum, erscheinungsdatum, ...) survives in anmerkung
        qual = "" if rolle in MENTION_ROLES else rolle
        flag = "erwaehnung" if rolle in MENTION_ROLES else dflag
        emit(out, src, block, "datum", dval, join_anm(anm, qual), flag)
        return

    # ---- place row (typ=ort) --------------------------------------------
    if typ == "ort":
        if rolle == "auffuehrungsort" or rolle in TRIGGER_ROLES:
            emit(out, src, block, "auffuehrungsort", name, anm)
        elif rolle in MENTION_ROLES:
            emit(out, src, block, "ort", name, anm, "erwaehnung")
        else:
            emit(out, src, block, "ort", name, join_anm(anm, rolle), "ort-rolle")
        return

    # ---- institution row -------------------------------------------------
    if typ == "institution":
        if rolle in EVENT_ROLES:
            emit(out, src, block, "ereignis", name, anm)
        elif rolle in MENTION_ROLES:
            emit(out, src, block, "organisation", name, anm, "vokab erwaehnung")
        else:
            emit(out, src, block, "organisation", name, anm, "vokab")
            emit_funktion(out, src, block, rolle)
        return

    # ---- person row ------------------------------------------------------
    if typ == "person":
        if rolle in MENTION_ROLES:
            emit(out, src, block, "person", name, anm, "erwaehnung")
        else:
            emit(out, src, block, "person", name, anm)
            emit_funktion(out, src, block, rolle)
        return

    # ---- ensemble row ----------------------------------------------------
    if typ == "ensemble":
        if rolle in MENTION_ROLES:
            emit(out, src, block, "ensemble", name, anm, "erwaehnung")
        else:
            emit(out, src, block, "ensemble", name, anm)
            emit_funktion(out, src, block, rolle)
        return

    # ---- ereignis row ----------------------------------------------------
    if typ == "ereignis":
        emit(out, src, block, "ereignis", name, anm)
        emit_funktion(out, src, block, rolle)
        return

    # ---- dokument row ----------------------------------------------------
    if typ == "dokument":
        if rolle in MENTION_ROLES:
            emit(out, src, block, "dokument", name, anm, "vokab erwaehnung")
        else:
            emit(out, src, block, "dokument", name, anm, "vokab")
            emit_funktion(out, src, block, rolle)
        return

    # ---- werk row --------------------------------------------------------
    if typ == "werk":
        flag = "erwaehnung" if rolle in MENTION_ROLES else (
            "aktivitaet-trigger" if rolle in TRIGGER_ROLES else "")
        emit(out, src, block, "werk", name, anm, flag)
        return

    # ---- fallback: unknown typ, carry verbatim + flag -------------------
    emit(out, src, block, typ or "?", name, join_anm(anm, rolle), "unmapped")


def emit_funktion(out, src, block, rolle):
    # mention and activity-trigger roles are not functions: skip here, they are
    # handled as entity flags / activity grouping respectively.
    if not rolle or rolle in MENTION_ROLES or rolle in TRIGGER_ROLES:
        return
    canon = FUNKTION_ALIASES.get(rolle, rolle)
    flag = "alias" if rolle in FUNKTION_ALIASES else (
        "" if canon in FUNKTION_ROLES else "vokab")
    emit(out, src, block, "funktion", canon, "", flag)


def join_anm(anm, qual):
    if qual and anm:
        return f"{anm} [{qual}]"
    if qual:
        return f"[{qual}]"
    return anm


def norm_date(raw):
    s = raw.strip()
    if ISO_DATE.match(s):
        return s, ""
    return s, "datum-format"  # ranges / free text: keep verbatim, flag


MONEY = re.compile(r"^\s*([\d][\d.,]*?)\s*[,\s]\s*(.+?)\s*$")


def split_money(raw):
    """Split 'AMOUNT, CURRENCY' verbatim; numeric normalization is deferred to
    the human/pipeline. Returns (amount_str, currency_str|"")."""
    m = MONEY.match(raw)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return raw.strip(), ""  # no currency token: keep amount verbatim


# --------------------------------------------------------------------------- #
# driver
# --------------------------------------------------------------------------- #
def migrate(prefix):
    src_rows = load_source(prefix)
    pidx = folio_person_index(src_rows)
    out = []
    for n, src in enumerate(src_rows):
        block = f"{src['folio']}#{n:04d}"
        transform_row(src, block, pidx, out)
    return src_rows, out


def write_outputs(prefix, out):
    """Flat CSV export (provenance preserved) for testing and diffing."""
    OUTDIR.mkdir(parents=True, exist_ok=True)
    safe = prefix.replace("/", "_")
    with open(OUTDIR / f"{safe}.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=OUT_COLS + ["_sheet"])
        w.writeheader()
        w.writerows(out)
    return safe


# --------------------------------------------------------------------------- #
# self-tests
# --------------------------------------------------------------------------- #
def run_tests(src_rows, out):
    results = []

    def check(name, ok, detail=""):
        results.append((name, ok, detail))

    # T1 provenance completeness: every source row produced >=1 output row
    src_provs = {r["prov"] for r in src_rows}
    out_provs = {r["_prov"] for r in out}
    missing = src_provs - out_provs
    check("T1 Vollstaendigkeit (jede Quellzeile -> >=1 Zielzeile)",
          not missing, f"{len(missing)} Quellzeilen ohne Output: {sorted(missing)[:5]}")

    # T2 soundness: every output row traces to an existing source row
    orphan = out_provs - src_provs
    check("T2 Korrektheit (jede Zielzeile -> existierende Quellzeile)",
          not orphan, f"{len(orphan)} verwaiste Zielzeilen")

    # T3 inverse / token conservation per source row: the union of output value
    # tokens reconstructs the source name tokens (minus pure-vocab funktion rows).
    by_prov = defaultdict(list)
    for r in out:
        by_prov[r["_prov"]].append(r)
    lost = []
    for s in src_rows:
        if not s["name"]:
            continue
        want = name_tokens(s["name"])
        got = set()
        for r in by_prov.get(s["prov"], []):
            if r["typ"] == "funktion":
                continue  # funktion comes from rolle, not name
            got |= name_tokens(r["value"])
        if not want <= got:
            lost.append((s["prov"], sorted(want - got)))
    check("T3 Token-Konservierung (Quellname vollstaendig im Output)",
          not lost, f"{len(lost)} Zeilen mit verlorenen Tokens: {lost[:5]}")

    # T4 inverse rolle/funktion: every source rolle that is a function appears
    # as a funktion value (canonical) somewhere in its block.
    missfun = []
    for s in src_rows:
        ro = lower(s["rolle"])
        canon = FUNKTION_ALIASES.get(ro, ro)
        if canon in FUNKTION_ROLES:
            vals = {r["value"] for r in by_prov.get(s["prov"], []) if r["typ"] == "funktion"}
            if canon not in vals:
                missfun.append((s["prov"], ro))
    check("T4 Funktion-Konservierung (Funktionsrollen -> funktion-Zeile)",
          not missfun, f"{len(missfun)} fehlende Funktionen: {missfun[:5]}")

    # T5 encoding guard: no replacement char in any output value/anmerkung
    bad = [r["_prov"] for r in out
           if has_replacement_char(r["value"]) or has_replacement_char(r["anmerkung"])]
    check("T5 Encoding-Guard (kein U+FFFD im Output)",
          not bad, f"{len(bad)} Zeilen mit Replacement-Char")

    # T6 aktivitaet_id stays empty everywhere
    filled = [r["_prov"] for r in out if r["aktivitaet_id"] != ""]
    check("T6 aktivitaet_id durchgehend leer",
          not filled, f"{len(filled)} Zeilen mit gesetzter ID")

    return results


def gold_7_29(src_rows, out):
    """Gold sample: no-loss at entity level for the Anleitung folio 7_29."""
    folio_out = [r for r in out if r["Folio"] == "7_29"]
    if not folio_out:
        return None
    values = {(r["typ"], r["value"]) for r in folio_out}
    valset = {r["value"] for r in folio_out}
    expect_activity = [
        ("ereignis", "Bayreuther Festspiele 1953"),
        ("werk", "Das Rheingold"),
        ("auffuehrungsort", "Bayreuth"),
        ("datum", "1953-07-25"),
    ]
    # every source person of the folio must survive as a person value
    src_persons = {s["name"] for s in src_rows
                   if s["folio"] == "7_29" and lower(s["typ"]) == "person"}
    miss_act = [e for e in expect_activity if e not in values]
    miss_pers = [p for p in src_persons if p not in valset]
    return {
        "activity_ok": not miss_act,
        "missing_activity": miss_act,
        "persons_ok": not miss_pers,
        "missing_persons": miss_pers,
        "n_out_rows": len(folio_out),
    }


def write_report(prefix, src_rows, out, results, gold):
    flags = defaultdict(int)
    for r in out:
        for f in r["_flag"].split():
            if f:
                flags[f] += 1
    lines = []
    lines.append(f"# Migration-Report v2 — {prefix}\n")
    lines.append(f"Quellzeilen: {len(src_rows)}  →  Zielzeilen: {len(out)}\n")
    lines.append("## Selbsttests\n")
    for name, ok, detail in results:
        mark = "PASS" if ok else "FAIL"
        lines.append(f"- [{mark}] {name}" + (f" — {detail}" if not ok else ""))
    lines.append("\n## Flag-Verteilung (Review-Hinweise)\n")
    for k in sorted(flags):
        lines.append(f"- {k}: {flags[k]}")
    if gold:
        lines.append("\n## Gold-Sample Folio 7_29 (Anleitungsbeispiel)\n")
        lines.append(f"- Zielzeilen: {gold['n_out_rows']}")
        lines.append(f"- Aktivitaets-Entitaeten vollstaendig: "
                     f"{'JA' if gold['activity_ok'] else 'NEIN ' + str(gold['missing_activity'])}")
        lines.append(f"- Alle Quell-Personen erhalten: "
                     f"{'JA' if gold['persons_ok'] else 'NEIN ' + str(gold['missing_persons'])}")
    (OUTDIR / "migration-report.md").write_text("\n".join(lines), encoding="utf-8")
    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# Vokabular-Glossar (Dropdown-Quelle + Definitionen + Ziel-Property)
# --------------------------------------------------------------------------- #
# ebene, term, definition, ziel-property, status. status=offen markiert noch
# zu bestaetigende Vokabular-Entscheidungen fuer den Operator.
VOKABULAR = [
    ("typ", "werk", "Musikalisches oder szenisches Werk (Oper, Lied), das aufgefuehrt, erwaehnt oder dokumentiert wird.", "rico:Record / m3gim:Work", "ok"),
    ("typ", "aktivitaet", "Markerzeile einer Aktivitaet; value benennt die Art (auffuehrung, gastspiel).", "m3gim:Performance", "ok"),
    ("typ", "ereignis", "Umgebender Rahmen einer Aktivitaet (Festival, Saison, Veranstaltung).", "m3gim:SpatiotemporalEvent", "ok"),
    ("typ", "auffuehrungsort", "Ort, an dem die Aktivitaet stattfindet.", "m3gim:performanceLocation", "ok"),
    ("typ", "ort", "Sonstiger Ort (Reise, Entstehung, Kontext); Subrolle in anmerkung.", "rico:Place", "ok"),
    ("typ", "datum", "Zeitangabe im Format JJJJ-MM-TT (oder JJJJ-MM, JJJJ).", "m3gim:DatedEvent", "ok"),
    ("typ", "person", "Einzelne mitwirkende oder genannte Person, Form 'Nachname, Vorname'.", "rico:Person", "ok"),
    ("typ", "funktion", "Aufgabe einer Person oder eines Ensembles in der Aktivitaet (kontrolliert).", "agrelon / m3gim:StageRole", "ok"),
    ("typ", "rolle", "Gesungene oder gespielte Partie (frei eingetragen).", "m3gim:StageRole", "ok"),
    ("typ", "gehalt", "Geldbetrag an eine Person; blanke Zahl ohne Tausenderpunkt.", "m3gim:monetaryAmount", "ok"),
    ("typ", "währung", "Waehrungskuerzel zum Betrag (eigene Zeile).", "m3gim:currency", "ok"),
    ("typ", "ensemble", "Kollektiv (Chor, Orchester), wirkt wie eine Person.", "rico:Group", "ok"),
    ("typ", "organisation", "Institution (Theater, Verlag, Firma, Sender).", "rico:CorporateBody", "offen"),
    ("typ", "einnahmen", "Finanzposten Einnahme auf Aktivitaetsebene (aus Altdaten).", "m3gim:DetailAnnotation", "offen"),
    ("typ", "ausgaben", "Finanzposten Ausgabe auf Aktivitaetsebene (aus Altdaten).", "m3gim:DetailAnnotation", "offen"),
    ("typ", "summe", "Finanzposten Summe auf Aktivitaetsebene (aus Altdaten).", "m3gim:DetailAnnotation", "offen"),
    ("typ", "dokument", "Verwiesenes Schriftstueck (Brief, Vertrag, Druck).", "rico:Record", "offen"),
    ("aktivitaet", "aufführung", "Oeffentliche Auffuehrung eines Werks.", "", "ok"),
    ("aktivitaet", "gastspiel", "Auswaertige Auffuehrung ausserhalb des Stammhauses.", "", "ok"),
    ("aktivitaet", "premiere", "Erstauffuehrung einer Produktion.", "", "ok"),
    ("aktivitaet", "probe", "Probe zu einer Produktion.", "", "ok"),
    ("aktivitaet", "aufnahme", "Ton- oder Rundfunkaufnahme.", "", "ok"),
    ("funktion", "sänger:in", "Singt eine Partie in der Auffuehrung.", "", "ok"),
    ("funktion", "dirigent:in", "Musikalische Leitung der Auffuehrung.", "", "ok"),
    ("funktion", "regisseur:in", "Szenische Leitung der Produktion.", "", "ok"),
    ("funktion", "komponist:in", "Komponiert das aufgefuehrte Werk.", "", "ok"),
    ("funktion", "chor", "Mitwirkendes Chor-Kollektiv.", "", "ok"),
    ("funktion", "chorleiter:in", "Einstudierung und Leitung des Chors.", "", "ok"),
    ("funktion", "choreograph:in", "Choreografie der Produktion.", "", "ok"),
    ("funktion", "ausstatter:in", "Buehnen- und Kostuemausstattung.", "", "ok"),
    ("funktion", "kostümbildner:in", "Entwurf der Kostueme.", "", "ok"),
    ("funktion", "maskenbildner:in", "Masken und Maskenbild.", "", "ok"),
    ("funktion", "beleuchter:in", "Beleuchtung der Auffuehrung.", "", "ok"),
    ("funktion", "technische leitung", "Technische Gesamtleitung.", "", "ok"),
    ("funktion", "bühnenleiter:in", "Leitung des Buehnenbetriebs.", "", "ok"),
    ("funktion", "repetitor:in", "Musikalische Einstudierung mit Solist:innen.", "", "ok"),
    ("funktion", "leitung", "Allgemeine Leitung, sofern nicht naeher bezeichnet.", "", "offen"),
    ("funktion", "interpret:in", "Ausfuehrende:r bei einer Aufnahme oder einem Konzert.", "", "ok"),
    ("funktion", "verfasser:in", "Verfasser:in eines Schriftstuecks (Brief, Vertrag).", "", "ok"),
    ("funktion", "herausgeber:in", "Herausgeber:in eines Programms oder Drucks.", "", "ok"),
    ("funktion", "agent:in", "Kuenstler:innenvermittlung im eigenen Auftrag.", "", "ok"),
    ("funktion", "vermittler:in", "Vermittelnde Stelle eines Engagements.", "", "ok"),
    ("funktion", "adressat:in", "Empfaenger:in eines Schriftstuecks.", "", "ok"),
    ("funktion", "veranstalter:in", "Veranstaltende Institution.", "", "ok"),
    ("funktion", "vertragspartner:in", "Partei eines Vertrags.", "", "ok"),
    ("funktion", "fluggesellschaft", "Befoerderndes Luftfahrtunternehmen (Reisebeleg).", "", "offen"),
]


def vokab_terms(ebene):
    return [t for e, t, *_ in VOKABULAR if e == ebene]


def build_workbook(out, path):
    """One Google-Sheets-ready workbook: Lies-mich, Vokabular, one sheet/box."""
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = openpyxl.Workbook()

    # --- Lies mich ---
    ws = wb.active
    ws.title = "Lies mich"
    readme = [
        ["M3GIM Verknuepfungen v2 — Arbeitsgrundlage", ""],
        ["", ""],
        ["Diese Mappe ist der gereinigte Altbestand im neuen Long-Format.", ""],
        ["Eine Zeile = eine Aussage. Pro Box ein Blatt.", ""],
        ["", ""],
        ["Spalte", "Bedeutung"],
        ["archivsignatur", "Signatur des Dokuments, je Zeile gleich."],
        ["Folio", "Blatt, auf dem die Information steht."],
        ["aktivitaet_id", "LEER. Wird von euch vergeben: Aktivitaet ganzzahlig (1), Beteiligung 1.01 ff. Erwaehnungen bleiben leer."],
        ["typ", "Art der Aussage, kontrolliert (siehe Blatt Vokabular)."],
        ["value", "Der Wert."],
        ["anmerkung", "Hinweise, Unsicherheiten, alte Subrollen in [eckigen Klammern]."],
        ["", ""],
        ["Hilfsspalte (vor dem finalen Pipeline-Lauf entfernbar):", ""],
        ["block", "Zeilen aus EINER Altzeile teilen denselben block; Hilfe beim Vergeben der aktivitaet_id pro Beteiligung."],
        ["", ""],
        ["Redundanz", "Besetzungen koennen mehrfach erscheinen (bare Partie, Komposit, Personenzeile). Das ist Absicht: bare Partie ohne Komposit = Saenger:in unklar. Beim Gruppieren zusammenfuehren."],
    ]
    for r in readme:
        ws.append(r)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95

    # --- Vokabular ---
    vk = wb.create_sheet("Vokabular")
    vk.append(["ebene", "term", "definition", "ziel-property", "status"])
    for row in VOKABULAR:
        vk.append(list(row))
    for col, w in zip("ABCDE", (12, 20, 70, 28, 8)):
        vk.column_dimensions[col].width = w

    # --- per box ---
    cols = ["archivsignatur", "Folio", "aktivitaet_id", "typ", "value",
            "anmerkung", "block"]
    by_sheet = defaultdict(list)
    for r in out:
        by_sheet[r["_sheet"]].append(r)
    for sheetname in sorted(by_sheet):
        safe_title = sheetname[:31]
        sh = wb.create_sheet(safe_title)
        sh.append(cols)
        for r in by_sheet[sheetname]:
            sh.append([
                r["archivsignatur"], r["Folio"], r["aktivitaet_id"], r["typ"],
                r["value"], r["anmerkung"], r["_block"],
            ])
        # aktivitaet_id (col C) as plain text -> no numeric coercion in Sheets
        for cell in sh["C"]:
            cell.number_format = "@"
        # typ dropdown (col D) from inline list (survives Sheets import)
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(vokab_terms("typ")),
                            allow_blank=True)
        sh.add_data_validation(dv)
        dv.add(f"D2:D{sh.max_row}")
    wb.save(path)
    return [s[:31] for s in sorted(by_sheet)]


def main():
    arg = sys.argv[1] if len(sys.argv) > 1 else "ALL"
    prefix = "" if arg.upper() == "ALL" else arg
    src_rows, out = migrate(prefix)
    safe = write_outputs(prefix or "ALL", out)
    results = run_tests(src_rows, out)
    gold = gold_7_29(src_rows, out)
    report = write_report(prefix or "ALL", src_rows, out, results, gold)
    print(report)
    # integration-ready workbook for Google Sheets upload
    wb_path = OUTDIR / "M3GIM-Verknuepfungen-v2.xlsx"
    sheets = build_workbook(out, wb_path)
    print(f"\nGeschrieben: data/migration/{safe}.csv (Flachexport)")
    print(f"Arbeitsmappe: data/migration/{wb_path.name}")
    print(f"  Blaetter: Lies mich, Vokabular, " + ", ".join(sheets))
    all_ok = all(ok for _, ok, _ in results)
    gold_ok = gold and gold["activity_ok"] and gold["persons_ok"]
    print(f"\nTests: {'ALLE PASS' if all_ok else 'FEHLER'} | "
          f"Gold 7_29: {'PASS' if gold_ok else 'FAIL'}")
    sys.exit(0 if all_ok and gold_ok else 1)


if __name__ == "__main__":
    main()
