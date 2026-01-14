#!/usr/bin/env python3
"""
M³GIM Migration Script
Transformiert Archiv-Export nach Google Sheets Excel-Format.
"""

import pandas as pd
import re
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

# Pfade
BASE_DIR = Path(__file__).parent.parent
ARCHIVE_DIR = BASE_DIR / "data" / "archive-export"
PROCESSED_DIR = BASE_DIR / "data" / "processed"

# Kontrollierte Vokabulare
VOCAB = {
    "dokumenttyp": ["autobiografie", "korrespondenz", "vertrag", "programm", "presse",
                    "repertoire", "studienunterlagen", "identitaetsdokument", "plakat",
                    "tontraeger", "sammlung"],
    "datierungsevidenz": ["aus_dokument", "erschlossen", "extern", "unbekannt"],
    "zugaenglichkeit": ["offen", "eingeschraenkt", "gesperrt"],
    "scan_status": ["nicht_gescannt", "gescannt", "online"],
    "sprache": ["de", "uk", "en", "fr", "it"],
    "fototyp": ["sw", "farbe", "digital"],
    "verknuepfung_typ": ["person", "ort", "institution", "werk", "ereignis"],
    "verknuepfung_rolle": ["vertragspartner", "interpretin", "verfasser", "adressat",
                          "erwaehnt", "unterzeichner", "vermittler", "veranstalter",
                          "entstehungsort", "zielort", "wohnort", "vertragsort",
                          "auffuehrungsort", "auftritt", "premiere", "rahmenveranstaltung", "implizit"]
}

# Keyword-Mapping für Dokumenttypen
DOKUMENTTYP_KEYWORDS = {
    "vertrag": ["vertrag", "engagement", "gage", "honorar"],
    "programm": ["programm", "programmheft", "programmzettel"],
    "presse": ["kritik", "rezension", "besprechung", "artikel"],
    "repertoire": ["repertoire", "rollenliste", "partien"],
    "autobiografie": ["lebenslauf", "biografie", "vita"],
    "identitaetsdokument": ["pass", "ausweis", "visum", "zeugnis", "diplom"],
    "studienunterlagen": ["studien", "noten", "partitur"]
}


def convert_date(val):
    """YYYYMMDD → YYYY-MM-DD"""
    if pd.isna(val):
        return ""
    try:
        s = str(int(val))
        if len(s) == 8:
            return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
        elif len(s) == 6:
            return f"{s[:4]}-{s[4:6]}"
        elif len(s) == 4:
            return s
    except (ValueError, TypeError):
        pass
    return ""


def convert_date_range(date_from, date_to):
    """Zwei Datumsfelder zu einem Zeitraum kombinieren"""
    d1 = convert_date(date_from)
    d2 = convert_date(date_to)
    if d1 and d2 and d1 != d2:
        return f"{d1}/{d2}"
    return d1 or d2


def map_dokumenttyp(systematik, enthaelt):
    """Systematikgruppe + Enthält → dokumenttyp"""
    if pd.isna(systematik):
        return "sammlung"

    systematik_lower = str(systematik).lower()

    # Direkte Mappings
    if "korrespondenz" in systematik_lower:
        return "korrespondenz"
    if "sammlung" in systematik_lower:
        return "sammlung"
    if "autobiograph" in systematik_lower:
        return "autobiografie"

    # Keyword-basiertes Mapping aus Enthält
    if not pd.isna(enthaelt):
        enthaelt_lower = str(enthaelt).lower()
        for typ, keywords in DOKUMENTTYP_KEYWORDS.items():
            if any(kw in enthaelt_lower for kw in keywords):
                return typ

    return "sammlung"


def map_zugaenglichkeit(sperrfrist, gesperrt_bis):
    """Sperrfrist → zugaenglichkeit"""
    if pd.isna(sperrfrist) and pd.isna(gesperrt_bis):
        return "offen"

    # Prüfe gesperrt_bis Datum
    if not pd.isna(gesperrt_bis):
        try:
            # Format: DD.MM.YYYY oder YYYY
            gesperrt_str = str(gesperrt_bis)
            if "." in gesperrt_str:
                parts = gesperrt_str.split(".")
                if len(parts) == 3:
                    year = int(parts[2])
                    if year > datetime.now().year:
                        return "gesperrt"
            return "offen"  # Datum in der Vergangenheit
        except (ValueError, TypeError):
            pass

    if not pd.isna(sperrfrist):
        return "eingeschraenkt"

    return "offen"


def map_scan_status(filename, speicherort):
    """Filename/Speicherort → scan_status"""
    if pd.isna(filename) and pd.isna(speicherort):
        return "nicht_gescannt"
    return "gescannt"


def map_fototyp(fototyp):
    """Fototyp-Mapping"""
    if pd.isna(fototyp):
        return "sw"
    fototyp_lower = str(fototyp).lower()
    if "farb" in fototyp_lower:
        return "farbe"
    if "pdf" in fototyp_lower or "digital" in fototyp_lower:
        return "digital"
    return "sw"


def split_protokoll(val):
    """Verzeichnungsprotokoll → (bearbeiter, erfassungsdatum)"""
    if pd.isna(val):
        return "", ""

    s = str(val)
    # Suche nach Datum im Format DD.MM.YYYY
    match = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", s)
    if match:
        day, month, year = match.groups()
        datum = f"{year}-{month}-{day}"
        bearbeiter = s[:match.start()].strip().rstrip("/")
        return bearbeiter, datum

    return s, ""


def format_excel(ws, dropdown_config):
    """Excel-Formatierung: Header, Spaltenbreiten, Dropdowns"""
    # Header-Formatierung
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Zeile einfrieren
    ws.freeze_panes = "A2"

    # Spaltenbreiten automatisch anpassen
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Dropdown-Validierung
    for col_name, values in dropdown_config.items():
        # Finde Spaltenindex
        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == col_name:
                col_idx = idx
                break

        if col_idx:
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(values)}"',
                allow_blank=True
            )
            dv.error = f"Bitte Wert aus Liste wählen"
            dv.errorTitle = "Ungültiger Wert"
            ws.add_data_validation(dv)
            # Auf alle Datenzeilen anwenden
            dv.add(f"{col_letter}2:{col_letter}{ws.max_row + 100}")


def migrate_objekte():
    """Migriert Hauptbestand, Plakate und Tonträger"""
    print("Migriere Objekte...")

    rows = []

    # Hauptbestand
    df_main = pd.read_excel(ARCHIVE_DIR / "Nachlass Malaniuk.xlsx")
    for _, row in df_main.iterrows():
        bearbeiter, erfassungsdatum = split_protokoll(row.get("Verzeichnungsprotokoll"))
        rows.append({
            "archivsignatur": row.get("Archivsignatur", ""),
            "box_nr": row.get("Box-Nr.", ""),
            "titel": row.get("Titel", ""),
            "entstehungsdatum": convert_date_range(row.get("Datierung von"), row.get("Datierung bis")),
            "datierungsevidenz": "",
            "dokumenttyp": map_dokumenttyp(row.get("Systematikgruppe 1"), row.get("Enthält")),
            "sprache": "",
            "umfang": row.get("Umfang", ""),
            "zugaenglichkeit": map_zugaenglichkeit(row.get("Sperrfrist"), row.get("gesperrt bis")),
            "scan_status": map_scan_status(row.get("Filename"), row.get("Speicherort")),
            "bearbeiter": bearbeiter,
            "erfassungsdatum": erfassungsdatum
        })
    print(f"  Hauptbestand: {len(df_main)} Einträge")

    # Plakate
    df_plakate = pd.read_excel(ARCHIVE_DIR / "Nachlass Malaniuk Plakate.xlsx")
    for _, row in df_plakate.iterrows():
        rows.append({
            "archivsignatur": row.get("Archivsignatur", ""),
            "box_nr": row.get("Mappen-Nr.", ""),
            "titel": row.get("Titel", ""),
            "entstehungsdatum": convert_date_range(row.get("Datierung von"), row.get("Datierung bis")),
            "datierungsevidenz": "",
            "dokumenttyp": "plakat",
            "sprache": "",
            "umfang": row.get("Format", ""),
            "zugaenglichkeit": "offen",
            "scan_status": "nicht_gescannt",
            "bearbeiter": "",
            "erfassungsdatum": ""
        })
    print(f"  Plakate: {len(df_plakate)} Einträge")

    # Tonträger
    df_ton = pd.read_excel(ARCHIVE_DIR / "Nachlass Malaniuk Tonträger.xlsx")
    for _, row in df_ton.iterrows():
        rows.append({
            "archivsignatur": row.get("Archivsignatur", ""),
            "box_nr": row.get("Box-Nr.", ""),
            "titel": row.get("Titel", ""),
            "entstehungsdatum": convert_date_range(row.get("Datierung von"), row.get("Datierung bis")),
            "datierungsevidenz": "",
            "dokumenttyp": "tontraeger",
            "sprache": "",
            "umfang": str(row.get("Dauer", "")),
            "zugaenglichkeit": "offen",
            "scan_status": "nicht_gescannt",
            "bearbeiter": "",
            "erfassungsdatum": ""
        })
    print(f"  Tonträger: {len(df_ton)} Einträge")

    # DataFrame erstellen und nach Signatur sortieren
    df_out = pd.DataFrame(rows)
    df_out = df_out.sort_values("archivsignatur")

    # Excel erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Objekte"

    for r_idx, row in enumerate(dataframe_to_rows(df_out, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Formatierung
    format_excel(ws, {
        "datierungsevidenz": VOCAB["datierungsevidenz"],
        "dokumenttyp": VOCAB["dokumenttyp"],
        "sprache": VOCAB["sprache"],
        "zugaenglichkeit": VOCAB["zugaenglichkeit"],
        "scan_status": VOCAB["scan_status"]
    })

    output_path = PROCESSED_DIR / "M3GIM-Objekte.xlsx"
    wb.save(output_path)
    print(f"  Gespeichert: {output_path} ({len(df_out)} Zeilen)")

    return df_out


def migrate_fotos():
    """Migriert Fotografien"""
    print("Migriere Fotos...")

    df_fotos = pd.read_excel(ARCHIVE_DIR / "Nachlass Malaniuk Fotos.xlsx")

    rows = []
    for _, row in df_fotos.iterrows():
        rows.append({
            "archivsignatur": row.get("Archivsignatur", ""),
            "alte_signatur": row.get("alte Archiv-Sign.", ""),
            "fotobox_nr": row.get("Fotobox-Nr.", ""),
            "titel": row.get("Titel", ""),
            "entstehungsdatum": convert_date_range(row.get("Datierung von"), row.get("Datierung bis")),
            "datierungsevidenz": "",
            "beschreibung": row.get("Beschreibung", ""),
            "stichwoerter": row.get("Stichwörter", ""),
            "fotograf": row.get("Fotograf", ""),
            "fototyp": map_fototyp(row.get("Fototyp")),
            "format": row.get("Format", ""),
            "aufnahmeort": row.get("Ort der Aufnahme", ""),
            "rechte": row.get("Rechte", ""),
            "filename": row.get("Filename", ""),
            "bearbeiter": "",
            "erfassungsdatum": ""
        })

    df_out = pd.DataFrame(rows)
    df_out = df_out.sort_values("archivsignatur")

    # Excel erstellen
    wb = Workbook()
    ws = wb.active
    ws.title = "Fotos"

    for r_idx, row in enumerate(dataframe_to_rows(df_out, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Formatierung
    format_excel(ws, {
        "datierungsevidenz": VOCAB["datierungsevidenz"],
        "fototyp": VOCAB["fototyp"]
    })

    output_path = PROCESSED_DIR / "M3GIM-Fotos.xlsx"
    wb.save(output_path)
    print(f"  Gespeichert: {output_path} ({len(df_out)} Zeilen)")

    return df_out


def main():
    """Hauptfunktion"""
    print("=" * 60)
    print("M³GIM Migration")
    print("=" * 60)

    # Output-Verzeichnis erstellen
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    # Migration durchführen
    df_objekte = migrate_objekte()
    df_fotos = migrate_fotos()

    print()
    print("=" * 60)
    print("Migration abgeschlossen")
    print(f"  Objekte: {len(df_objekte)}")
    print(f"  Fotos: {len(df_fotos)}")
    print(f"  Ausgabe: {PROCESSED_DIR}")
    print("=" * 60)


if __name__ == "__main__":
    main()
