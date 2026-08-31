"""Explorativer Diff der M3GIM-Verknuepfungstabelle ALT gegen NEU.

Read-only. Laedt beide Mappen mit derselben Logik wie
``scripts/transform.py::load_verknuepfungen`` (Spalte 0 positionell als
Archivsignatur, ffill je Blatt, Blaetter ohne ``typ``+``name`` uebersprungen)
und schreibt einen Markdown-Bericht.

Aufruf:
    python diff_verknuepfungen.py ALT.xlsx NEU.xlsx OBJEKTE.xlsx \
        --out reports/diff-verknuepfungen.md [--csv-box1 "... Box 1.csv"]
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding="utf-8")


# ---------------------------------------------------------------------------
# Ladelogik, deckungsgleich mit transform.load_verknuepfungen
# ---------------------------------------------------------------------------

def normalize_signatur(sig: str) -> str:
    """NIM-Konvolutnummer auf drei Stellen nullen (NIM_11 -> NIM_011)."""
    return re.sub(r"NIM_(\d{1,3})\b", lambda m: f"NIM_{int(m.group(1)):03d}", sig)


def load_verknuepfungen(path: Path) -> tuple[pd.DataFrame, list[str]]:
    """Alle Verknuepfungsblaetter einer Mappe zu einer DataFrame."""
    xl = pd.ExcelFile(path)
    frames: list[pd.DataFrame] = []
    skipped: list[str] = []

    for sheet in xl.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet)
        if df.empty:
            skipped.append(f"{sheet} (leer)")
            continue

        rename: dict = {}
        for pos, col in enumerate(df.columns):
            if pos == 0:
                rename[col] = "archivsignatur"
            elif isinstance(col, str):
                rename[col] = col.strip().lower()
        df = df.rename(columns=rename)

        if not {"typ", "name"}.issubset(df.columns):
            skipped.append(f"{sheet} (keine typ/name-Spalten)")
            continue

        df["_orig_columns"] = ", ".join(
            str(c) for c in pd.read_excel(path, sheet_name=sheet, nrows=0).columns
        )
        df["archivsignatur"] = df["archivsignatur"].ffill()
        df["archivsignatur"] = df["archivsignatur"].map(
            lambda s: normalize_signatur(s) if isinstance(s, str) else s)
        df["_xlsx_sheet"] = sheet
        df["_xlsx_row"] = [int(i) + 2 for i in range(len(df))]
        frames.append(df)

    if not frames:
        return pd.DataFrame(columns=["archivsignatur", "_xlsx_sheet", "_xlsx_row"]), skipped
    return pd.concat(frames, ignore_index=True, sort=False), skipped


# ---------------------------------------------------------------------------
# Normalisierungshelfer
# ---------------------------------------------------------------------------

def cell(value) -> str:
    """Zellwert als getrimmter String, NaN als Leerstring."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if pd.isna(value):
        return ""
    s = str(value).strip()
    return "" if s.lower() == "nan" else s


def norm_folio(value) -> str:
    """Folio-Wert vergleichbar machen: Excel-Zeitanteil ab, Float-Ganzzahl zu int."""
    s = cell(value)
    if not s:
        return ""
    s = re.sub(r"\s+00:00:00$", "", s)
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def norm_role(value) -> str:
    """Rollenwert wie transform.normalize_role: lower, strip, Gender-Suffix ab."""
    v = cell(value).lower()
    for suffix in (":innen", ":in"):
        if v.endswith(suffix):
            return v[: -len(suffix)]
    if v.endswith("*innen"):
        return v[:-6]
    if v.endswith("*in"):
        return v[:-3]
    return v


def decompose_typ(typ: str) -> tuple[str, ...]:
    """Komposit-Typ zerlegen wie transform.decompose_komposit_typ."""
    parts = [t.strip().lower() for t in re.split(r"[,_]", typ)]
    if (len(parts) == 2 and parts[0] == "rolle"
            and "nger" in parts[1] and parts[1].split()[:2] == ["vorname", "nachname"]):
        return ("rolle", "person")
    return tuple(p for p in parts if p not in ("waehrung", "währung"))


# ---------------------------------------------------------------------------
# Muster-Klassifikation
# ---------------------------------------------------------------------------

def folio_pattern(value) -> str:
    s = cell(value)
    if not s:
        return "leer"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2} 00:00:00", s):
        return "datum-autokonvertiert (YYYY-MM-DD 00:00:00)"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", s):
        return "datum (ISO)"
    if re.fullmatch(r"\d+\.0", s) or re.fullmatch(r"\d+", s):
        return "int"
    if re.fullmatch(r"\d+_\d+", s):
        return "n_m"
    if re.fullmatch(r"\d+_\d+_\d+", s):
        return "n_m_k"
    if s.lower() == "folio":
        return "Literal 'Folio'"
    return f"sonstiges ({s})"


_ISO_FULL = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_ISO_MONTH = re.compile(r"^\d{4}-\d{2}$")
_ISO_YEAR = re.compile(r"^\d{4}$")
_ISO_SPAN = re.compile(r"^\d{4}(-\d{2}(-\d{2})?)?/\d{4}(-\d{2}(-\d{2})?)?$")
_QUALIFIER = re.compile(r"^(circa|vor|nach|ca)\s*:", re.IGNORECASE)
_BIS = re.compile(r"\bbis\b", re.IGNORECASE)
_TIMESTAMP = re.compile(r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$")


def date_pattern(value) -> str:
    s = cell(value)
    if not s:
        return "leer"
    if _TIMESTAMP.match(s):
        return "Timestamp-Autokonvertierung (YYYY-MM-DD 00:00:00)"
    if _QUALIFIER.match(s):
        return "Qualifier (circa/vor/nach)"
    if _BIS.search(s):
        return "Zeitspanne mit 'bis'"
    if _ISO_FULL.match(s):
        return "ISO YYYY-MM-DD"
    if _ISO_MONTH.match(s):
        return "ISO YYYY-MM"
    if _ISO_YEAR.match(s):
        return "ISO YYYY"
    if _ISO_SPAN.match(s):
        return "ISO-Zeitspanne mit /"
    if re.match(r"^\d{4}-\d{4}$", s):
        return "Jahresspanne mit Bindestrich"
    if re.match(r"^\d{1,2}\.\d{1,2}\.\d{4}$", s):
        return "deutsches Datum (DD.MM.YYYY)"
    return "Freitext / sonstiges"


# ---------------------------------------------------------------------------
# Berichtsbausteine
# ---------------------------------------------------------------------------

def sheet_signatures(df: pd.DataFrame) -> dict[str, set[str]]:
    out: dict[str, set[str]] = defaultdict(set)
    for sheet, sig in zip(df["_xlsx_sheet"], df["archivsignatur"]):
        s = cell(sig)
        if s:
            out[sheet].add(s)
    return out


def strip_excel_time(value) -> str:
    """Excel-Zeitanteil '00:00:00' abstreifen, wie transform.clean_date."""
    return re.sub(r"\s+00:00:00$", "", cell(value))


def row_key(row) -> tuple:
    """Strikter Schluessel: Rohwerte, nur typ kleingeschrieben."""
    return (
        cell(row.get("archivsignatur")),
        norm_folio(row.get("folio")),
        cell(row.get("typ")).lower(),
        cell(row.get("name")),
    )


def row_key_norm(row) -> tuple:
    """Normalisierter Schluessel: Komposit-Typ zerlegt, Excel-Zeitanteil ab.

    Faengt die beiden Formatwechsel des neuen Exports ab, die sonst jede
    betroffene Zeile als geloescht plus neu erscheinen lassen: die Umbenennung
    ``ausgaben, waehrung`` -> ``ausgaben_waehrung`` und die Autokonvertierung
    eines ISO-Datums zum Timestamp.
    """
    return (
        cell(row.get("archivsignatur")),
        norm_folio(row.get("folio")),
        "+".join(decompose_typ(cell(row.get("typ")))),
        strip_excel_time(row.get("name")),
    )


def row_payload(row) -> tuple:
    return (norm_role(row.get("rolle")), cell(row.get("anmerkung")))


def build_key_map(df: pd.DataFrame, keyfn=row_key) -> dict[tuple, list[dict]]:
    out: dict[tuple, list[dict]] = defaultdict(list)
    for _, row in df.iterrows():
        out[keyfn(row)].append({
            "payload": row_payload(row),
            "sheet": cell(row.get("_xlsx_sheet")),
            "row": cell(row.get("_xlsx_row")),
        })
    return out


def load_typ_rolle(path: Path) -> dict[str, list[str]] | None:
    """Blatt Typ-Rolle als {typ: [rollen]} laden, falls vorhanden."""
    xl = pd.ExcelFile(path)
    sheet = next((s for s in xl.sheet_names if s.strip().lower() in
                  ("typ-rolle", "typ-rollen")), None)
    if sheet is None:
        return None
    df = pd.read_excel(path, sheet_name=sheet, header=None)
    mapping: dict[str, list[str]] = {}
    for _, row in df.iloc[1:].iterrows():
        vals = [cell(v) for v in row.tolist()]
        typ = vals[0] if vals else ""
        rollen = [v for v in vals[1:] if v]
        if typ:
            mapping[typ] = rollen
    return mapping


def load_objekt_ids(path: Path) -> tuple[set[str], set[str], str]:
    """Objekt-IDs (signatur + ' ' + folio) und reine Signaturen der Objekttabelle."""
    df = pd.read_excel(path)
    df.columns = [c.lower().strip() if isinstance(c, str) else c for c in df.columns]
    folio_col = None
    for col in df.columns:
        if not isinstance(col, str):
            continue
        if col.lower() in ("folio", "folio nr", "folio_nr") or "unnamed" in col.lower():
            sample = df[col].dropna().astype(str).head(5)
            if any(re.match(r"^\d+_\d+$", s.strip()) or s.strip().startswith("fol.")
                   for s in sample):
                folio_col = col
                break
    ids: set[str] = set()
    sigs: set[str] = set()
    for _, row in df.iterrows():
        sig = cell(row.get("archivsignatur"))
        if not sig:
            continue
        sig = normalize_signatur(sig)
        sigs.add(sig)
        folio = norm_folio(row.get(folio_col)) if folio_col else ""
        ids.add(f"{sig} {folio}" if folio else sig)
    return ids, sigs, str(folio_col)


def counter_table(alt: Counter, neu: Counter, head: str) -> list[str]:
    keys = sorted(set(alt) | set(neu), key=lambda k: (-neu.get(k, 0), -alt.get(k, 0), k))
    lines = [f"| {head} | ALT | NEU | Delta |", "|---|---:|---:|---:|"]
    for k in keys:
        a, n = alt.get(k, 0), neu.get(k, 0)
        lines.append(f"| `{k}` | {a} | {n} | {n - a:+d} |")
    return lines


# ---------------------------------------------------------------------------
# Hauptlauf
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("alt", type=Path)
    ap.add_argument("neu", type=Path)
    ap.add_argument("objekte", type=Path)
    ap.add_argument("--out", type=Path, default=Path("reports/diff-verknuepfungen.md"))
    ap.add_argument("--csv-box1", type=Path, default=None)
    ap.add_argument("--zweitfassung", type=Path, default=None,
                    help="weitere Mappe, gegen ALT geprueft (Abschnitt 12)")
    ap.add_argument("--sample", type=int, default=25,
                    help="max. Beispielzeilen je Diff-Block")
    args = ap.parse_args()

    df_alt, skip_alt = load_verknuepfungen(args.alt)
    df_neu, skip_neu = load_verknuepfungen(args.neu)

    L: list[str] = []
    add = L.append

    add("# Diff der Verknuepfungstabelle ALT gegen NEU")
    add("")
    add(f"ALT `{args.alt}`  ")
    add(f"NEU `{args.neu}`  ")
    add(f"Objekttabelle `{args.objekte}`")
    add("")
    add(f"Geladene Zeilen ALT {len(df_alt)}, NEU {len(df_neu)}.")
    add("")
    add(f"Uebersprungene Blaetter ALT: {skip_alt or 'keine'}")
    add("")
    add(f"Uebersprungene Blaetter NEU: {skip_neu or 'keine'}")
    add("")

    # --- 1 Blattzuordnung -------------------------------------------------
    add("## 1 Blattzuordnung ueber Signaturmengen")
    add("")
    sig_alt = sheet_signatures(df_alt)
    sig_neu = sheet_signatures(df_neu)
    add("| Blatt ALT | Zeilen | bestes Blatt NEU | Jaccard | Signaturen gemeinsam / nur ALT / nur NEU |")
    add("|---|---:|---|---:|---|")
    matched_neu: set[str] = set()
    for sa, sigs_a in sig_alt.items():
        best, best_j, best_stats = None, 0.0, (0, 0, 0)
        for sn, sigs_n in sig_neu.items():
            inter = len(sigs_a & sigs_n)
            union = len(sigs_a | sigs_n)
            j = inter / union if union else 0.0
            if j > best_j:
                best, best_j = sn, j
                best_stats = (inter, len(sigs_a - sigs_n), len(sigs_n - sigs_a))
        if best:
            matched_neu.add(best)
        rows_a = int((df_alt["_xlsx_sheet"] == sa).sum())
        add(f"| {sa} | {rows_a} | {best or '(keins)'} | {best_j:.3f} | "
            f"{best_stats[0]} / {best_stats[1]} / {best_stats[2]} |")
    add("")
    unmatched = [s for s in sig_neu if s not in matched_neu]
    add(f"Blaetter NEU ohne ALT-Gegenstueck: {unmatched or 'keine'}")
    add("")

    # --- 2 Zeilenzahl + Spalten ------------------------------------------
    add("## 2 Zeilenzahl und Spaltenkoepfe je Blatt")
    add("")
    add("| Mappe | Blatt | Zeilen | Original-Spaltenkoepfe |")
    add("|---|---|---:|---|")
    for label, df in (("ALT", df_alt), ("NEU", df_neu)):
        for sheet in df["_xlsx_sheet"].dropna().unique():
            sub = df[df["_xlsx_sheet"] == sheet]
            cols = sub["_orig_columns"].iloc[0] if "_orig_columns" in sub else ""
            add(f"| {label} | {sheet} | {len(sub)} | `{cols}` |")
    add("")
    add("### Spalte datenpunkt_id gegen data_id")
    add("")
    add("| Mappe | Blatt | datenpunkt_id | data_id |")
    add("|---|---|---|---|")
    for label, df in (("ALT", df_alt), ("NEU", df_neu)):
        for sheet in df["_xlsx_sheet"].dropna().unique():
            sub = df[df["_xlsx_sheet"] == sheet]
            cols = str(sub["_orig_columns"].iloc[0]).lower()
            add(f"| {label} | {sheet} | {'ja' if 'datenpunkt_id' in cols else 'nein'} "
                f"| {'ja' if re.search(r'(^|, )data_id', cols) else 'nein'} |")
    add("")

    # --- 3 Werteverteilung typ / rolle -----------------------------------
    add("## 3 Werteverteilung typ (exakte Schreibweise)")
    add("")
    ct_alt = Counter(cell(v) for v in df_alt.get("typ", pd.Series(dtype=object)) if cell(v))
    ct_neu = Counter(cell(v) for v in df_neu.get("typ", pd.Series(dtype=object)) if cell(v))
    L.extend(counter_table(ct_alt, ct_neu, "typ"))
    add("")

    add("## 4 Werteverteilung rolle (exakte Schreibweise)")
    add("")
    cr_alt = Counter(cell(v) for v in df_alt.get("rolle", pd.Series(dtype=object)) if cell(v))
    cr_neu = Counter(cell(v) for v in df_neu.get("rolle", pd.Series(dtype=object)) if cell(v))
    L.extend(counter_table(cr_alt, cr_neu, "rolle"))
    add("")

    add("### Rollen nach Genderform (NEU)")
    add("")
    forms = defaultdict(list)
    for r in sorted(set(cr_neu)):
        if ":innen" in r or r.endswith(":in"):
            forms["Doppelpunkt-Form"].append(r)
        elif "*in" in r:
            forms["Sternchen-Form"].append(r)
        elif re.search(r"[a-zäöü]In(nen)?\b", r):
            forms["Binnen-I"].append(r)
        else:
            forms["ungegendert"].append(r)
    for k in ("Doppelpunkt-Form", "Sternchen-Form", "Binnen-I", "ungegendert"):
        vals = forms.get(k, [])
        add(f"**{k}** ({len(vals)}): " + (", ".join(f"`{v}`" for v in vals) or "keine"))
        add("")

    # --- 5 Kreuztabelle typ x rolle gegen Typ-Rolle -----------------------
    add("## 5 Kreuztabelle typ x rolle NEU gegen das Blatt Typ-Rolle")
    add("")
    tr = load_typ_rolle(args.neu)
    if tr is None:
        add("Kein Blatt `Typ-Rolle` in der NEU-Mappe gefunden.")
        add("")
    else:
        add("Blatt `Typ-Rolle`, wortgetreu:")
        add("")
        add("| Typ | Rollen |")
        add("|---|---|")
        for t, rs in tr.items():
            add(f"| `{t}` | " + (", ".join(f"`{r}`" for r in rs) or "(leer)") + " |")
        add("")

        allowed: dict[tuple[str, ...], set[str]] = {}
        for t, rs in tr.items():
            allowed.setdefault(decompose_typ(t), set()).update(norm_role(r) for r in rs)

        observed: Counter = Counter()
        raw_by_pair: dict[tuple, Counter] = defaultdict(Counter)
        for _, row in df_neu.iterrows():
            t = cell(row.get("typ"))
            if not t:
                continue
            key = decompose_typ(t)
            r = norm_role(row.get("rolle"))
            observed[(key, r)] += 1
            raw_by_pair[(key, r)][(t, cell(row.get("rolle")))] += 1

        add("### Kombinationen in den Daten, die Typ-Rolle nicht kennt")
        add("")
        add("| typ (normalisiert) | rolle (normalisiert) | Zeilen | Rohformen (typ / rolle) |")
        add("|---|---|---:|---|")
        gaps = 0
        for (key, r), n in sorted(observed.items(), key=lambda kv: -kv[1]):
            if key not in allowed:
                note = "typ fehlt im Blatt"
            elif r and r not in allowed[key]:
                note = "rolle fehlt beim typ"
            else:
                continue
            gaps += 1
            raws = "; ".join(f"`{a}` / `{b or ''}`"
                             for (a, b), _ in raw_by_pair[(key, r)].most_common(3))
            add(f"| `{'‚ '.join(key) if key else '(leer)'}` | `{r or '(leer)'}` | {n} | {raws} ({note}) |")
        if gaps == 0:
            add("| (keine) | | | |")
        add("")

        add("### Kombinationen aus Typ-Rolle ohne Beleg in den Daten")
        add("")
        add("| Typ (Blatt) | Rolle (Blatt) |")
        add("|---|---|")
        seen = {(k, r) for (k, r) in observed}
        unused = 0
        for t, rs in tr.items():
            key = decompose_typ(t)
            for r in rs:
                if (key, norm_role(r)) not in seen:
                    unused += 1
                    add(f"| `{t}` | `{r}` |")
        if unused == 0:
            add("| (keine) | |")
        add("")

    # --- 6 Fuellgrad datenpunkt_id / data_id -----------------------------
    add("## 6 Fuellgrad der Buendelungsspalte je Blatt")
    add("")
    add("| Mappe | Blatt | Spaltenname | Zeilen | gefuellt | Anteil | Wertebeispiele |")
    add("|---|---|---|---:|---:|---:|---|")
    for label, df in (("ALT", df_alt), ("NEU", df_neu)):
        for sheet in df["_xlsx_sheet"].dropna().unique():
            sub = df[df["_xlsx_sheet"] == sheet]
            colname = None
            for cand in ("datenpunkt_id", "data_id"):
                if cand in sub.columns and sub[cand].notna().any():
                    colname = cand
                    break
            if colname is None:
                cols = str(sub["_orig_columns"].iloc[0]).lower()
                colname = "datenpunkt_id" if "datenpunkt_id" in cols else (
                    "data_id" if "data_id" in cols else "(fehlt)")
                filled = 0
                sample = ""
            else:
                vals = [cell(v) for v in sub[colname] if cell(v)]
                filled = len(vals)
                sample = ", ".join(sorted(set(vals), key=str)[:6])
            add(f"| {label} | {sheet} | {colname} | {len(sub)} | {filled} | "
                f"{filled / max(len(sub), 1):.1%} | {sample} |")
    add("")

    add("### Zellwert und Zahlenformat der Buendelungsspalte (NEU)")
    add("")
    add("Der gelesene Wert allein sagt nicht, was erfasst wurde. Ein Datumsformat "
        "in dieser Spalte belegt, dass die Tabellenkalkulation eine Eingabe der "
        "Form `1.01` als Kalenderdatum gelesen hat.")
    add("")
    add("| Blatt | Zellwert | Anzahl | Zahlenformat |")
    add("|---|---|---:|---|")
    wb = openpyxl.load_workbook(args.neu, data_only=True, read_only=True)
    for sheet in df_neu["_xlsx_sheet"].dropna().unique():
        ws = wb[sheet]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = [i for i, h in enumerate(header)
               if str(h).strip().lower() in ("datenpunkt_id", "data_id")]
        if not idx:
            continue
        seen: Counter = Counter()
        for r in ws.iter_rows(min_row=2):
            if idx[0] >= len(r):
                continue
            c = r[idx[0]]
            if c.value is None:
                continue
            seen[(f"{type(c.value).__name__}: {c.value}", c.number_format)] += 1
        for (val, fmt), n in seen.most_common():
            add(f"| {sheet} | `{val}` | {n} | `{fmt}` |")
    wb.close()
    add("")

    # --- 7 Folio-Muster ---------------------------------------------------
    add("## 7 Folio-Muster")
    add("")
    fp_alt = Counter(folio_pattern(v) for v in df_alt.get("folio", pd.Series(dtype=object)))
    fp_neu = Counter(folio_pattern(v) for v in df_neu.get("folio", pd.Series(dtype=object)))
    # sonstiges-Werte buendeln
    def bundle(c: Counter) -> Counter:
        out = Counter()
        for k, v in c.items():
            out["sonstiges" if k.startswith("sonstiges") else k] += v
        return out
    L.extend(counter_table(bundle(fp_alt), bundle(fp_neu), "Folio-Muster"))
    add("")
    add("Belegte Nicht-Standard-Folios NEU (Datum, Literal, sonstiges), mit Fundstelle:")
    add("")
    add("| Blatt | XLSX-Zeile | Signatur | Folio-Rohwert | Muster |")
    add("|---|---:|---|---|---|")
    shown = 0
    for _, row in df_neu.iterrows():
        pat = folio_pattern(row.get("folio"))
        if pat in ("int", "n_m", "n_m_k", "leer"):
            continue
        shown += 1
        if shown > args.sample:
            continue
        add(f"| {cell(row.get('_xlsx_sheet'))} | {cell(row.get('_xlsx_row'))} | "
            f"{cell(row.get('archivsignatur'))} | `{cell(row.get('folio'))}` | {pat} |")
    if shown == 0:
        add("| (keine) | | | | |")
    elif shown > args.sample:
        add(f"| ... | | | | insgesamt {shown} Zeilen |")
    add("")

    # --- 8 Datumsformate --------------------------------------------------
    add("## 8 Datumsformate in name bei typ Datum")
    add("")
    def date_counter(df: pd.DataFrame) -> Counter:
        c = Counter()
        for _, row in df.iterrows():
            t = cell(row.get("typ")).lower()
            if "datum" not in t:
                continue
            c[date_pattern(row.get("name"))] += 1
        return c
    L.extend(counter_table(date_counter(df_alt), date_counter(df_neu), "Datumsmuster"))
    add("")
    add("Nicht-ISO-Datumswerte NEU (Freitext, Qualifier, bis-Spanne, Timestamp), Beispiele:")
    add("")
    add("| Blatt | XLSX-Zeile | Signatur | typ | name-Rohwert | Muster |")
    add("|---|---:|---|---|---|---|")
    shown = 0
    for _, row in df_neu.iterrows():
        t = cell(row.get("typ")).lower()
        if "datum" not in t:
            continue
        pat = date_pattern(row.get("name"))
        if pat.startswith("ISO") or pat == "leer":
            continue
        shown += 1
        if shown > args.sample:
            continue
        add(f"| {cell(row.get('_xlsx_sheet'))} | {cell(row.get('_xlsx_row'))} | "
            f"{cell(row.get('archivsignatur'))} | `{cell(row.get('typ'))}` | "
            f"`{cell(row.get('name'))}` | {pat} |")
    if shown == 0:
        add("| (keine) | | | | | |")
    elif shown > args.sample:
        add(f"| ... | | | | | insgesamt {shown} Zeilen |")
    add("")

    # --- 9 Zeilendiff -----------------------------------------------------
    add("## 9 Zeilendiff (Schluessel Signatur + Folio + typ + name)")
    add("")
    strict_alt = build_key_map(df_alt, row_key)
    strict_neu = build_key_map(df_neu, row_key)
    map_alt = build_key_map(df_alt, row_key_norm)
    map_neu = build_key_map(df_neu, row_key_norm)
    only_neu = [k for k in map_neu if k not in map_alt]
    only_alt = [k for k in map_alt if k not in map_neu]
    changed = []
    for k in map_neu:
        if k not in map_alt:
            continue
        pa = {e["payload"] for e in map_alt[k]}
        pn = {e["payload"] for e in map_neu[k]}
        if pa != pn:
            changed.append((k, pa, pn))
    add("Zwei Schluesselvarianten. Der strikte Schluessel nimmt die Rohwerte, der "
        "normalisierte zerlegt den Komposit-Typ und streift den Excel-Zeitanteil "
        "des name-Werts ab. Die Differenz beider Zahlen ist reiner Formatwechsel "
        "des Exports, keine inhaltliche Aenderung.")
    add("")
    add("| Schluessel | nur NEU | nur ALT |")
    add("|---|---:|---:|")
    add(f"| strikt | {len([k for k in strict_neu if k not in strict_alt])} "
        f"| {len([k for k in strict_alt if k not in strict_neu])} |")
    add(f"| normalisiert | {len(only_neu)} | {len(only_alt)} |")
    add("")
    add(f"Normalisiert: Schluessel nur NEU {len(only_neu)}, nur ALT {len(only_alt)}, "
        f"geaendert (rolle/anmerkung) {len(changed)}.")
    add("")
    add(f"Zeilen hinter den Nur-NEU-Schluesseln: "
        f"{sum(len(map_neu[k]) for k in only_neu)}; hinter den Nur-ALT-Schluesseln: "
        f"{sum(len(map_alt[k]) for k in only_alt)}.")
    add("")

    add("### Neue Signaturen (in NEU, nicht in ALT)")
    add("")
    sigs_a = {k[0] for k in map_alt}
    sigs_n = {k[0] for k in map_neu}
    add("Nur NEU: " + (", ".join(f"`{s}`" for s in sorted(sigs_n - sigs_a)) or "keine"))
    add("")
    add("Nur ALT: " + (", ".join(f"`{s}`" for s in sorted(sigs_a - sigs_n)) or "keine"))
    add("")

    add("### Nur-NEU-Zeilen je Blatt")
    add("")
    per_sheet = Counter()
    for k in only_neu:
        for e in map_neu[k]:
            per_sheet[e["sheet"]] += 1
    add("| Blatt NEU | neue Zeilen |")
    add("|---|---:|")
    for s, n in per_sheet.most_common():
        add(f"| {s} | {n} |")
    add("")

    add("### Nur-ALT-Zeilen (entfallen), Beispiele")
    add("")
    add("| Blatt ALT | XLSX-Zeile | Signatur | Folio | typ | name |")
    add("|---|---:|---|---|---|---|")
    for i, k in enumerate(sorted(only_alt)):
        if i >= args.sample:
            add(f"| ... | | | | | insgesamt {len(only_alt)} Schluessel |")
            break
        e = map_alt[k][0]
        add(f"| {e['sheet']} | {e['row']} | {k[0]} | `{k[1]}` | `{k[2]}` | `{k[3][:70]}` |")
    if not only_alt:
        add("| (keine) | | | | | |")
    add("")

    add("### Geaenderte Zeilen (gleicher Schluessel, andere rolle oder anmerkung), Beispiele")
    add("")
    add("| Signatur | Folio | typ | name | ALT (rolle, anmerkung) | NEU (rolle, anmerkung) |")
    add("|---|---|---|---|---|---|")
    for i, (k, pa, pn) in enumerate(sorted(changed)):
        if i >= args.sample:
            add(f"| ... | | | | | insgesamt {len(changed)} Schluessel |")
            break
        add(f"| {k[0]} | `{k[1]}` | `{k[2]}` | `{k[3][:50]}` | "
            f"{sorted(pa)} | {sorted(pn)} |".replace("\n", " "))
    if not changed:
        add("| (keine) | | | | | |")
    add("")

    # --- 10 Referenzielle Pruefung ---------------------------------------
    add("## 10 Referenzielle Pruefung Signatur + Folio gegen die Objekttabelle")
    add("")
    obj_ids, obj_sigs, folio_col = load_objekt_ids(args.objekte)
    add(f"Objekttabelle: {len(obj_ids)} Objekt-IDs, {len(obj_sigs)} Signaturen, "
        f"Folio-Spalte `{folio_col}`.")
    add("")
    def check_refs(df: pd.DataFrame):
        missing: Counter = Counter()
        missing_sig: Counter = Counter()
        example: dict[str, tuple] = {}
        for _, row in df.iterrows():
            sig = cell(row.get("archivsignatur"))
            if not sig or sig.lower() == "beispiel":
                continue
            folio = norm_folio(row.get("folio"))
            if folio.lower() == "folio":
                folio = ""
            oid = f"{sig} {folio}" if folio else sig
            if oid in obj_ids:
                continue
            missing[oid] += 1
            if sig not in obj_sigs:
                missing_sig[sig] += 1
            example.setdefault(oid, (cell(row.get("_xlsx_sheet")),
                                     cell(row.get("_xlsx_row"))))
        return missing, missing_sig, example

    missing_alt, _, _ = check_refs(df_alt)
    missing, missing_sig, example = check_refs(df_neu)
    add(f"Zum Vergleich gegen dieselbe (neue) Objekttabelle: ALT "
        f"{sum(missing_alt.values())} Zeilen auf {len(missing_alt)} Objekt-IDs ohne "
        f"Objektsatz.")
    add("")
    add(f"NEU: Verknuepfungszeilen ohne passenden Objektsatz {sum(missing.values())} "
        f"auf {len(missing)} Objekt-IDs.")
    add("")
    add(f"Davon Signaturen, die die Objekttabelle gar nicht kennt: "
        f"{sum(missing_sig.values())} Zeilen auf {len(missing_sig)} Signaturen "
        f"({', '.join(f'`{s}`' for s in sorted(missing_sig)) or 'keine'}).")
    add("")
    add("| Objekt-ID ohne Objektsatz | Zeilen NEU | Zeilen ALT | Blatt | erste XLSX-Zeile |")
    add("|---|---:|---:|---|---:|")
    for oid, n in missing.most_common():
        sheet, rrow = example[oid]
        add(f"| `{oid}` | {n} | {missing_alt.get(oid, 0)} | {sheet} | {rrow} |")
    if not missing:
        add("| (keine) | | | | |")
    add("")
    gone = [oid for oid in missing_alt if oid not in missing]
    add("Objekt-IDs, die nur in ALT unaufgeloest waren. Sie sind entweder "
        "quellseitig geklaert oder tragen in NEU eine andere Schreibform und "
        "erscheinen dann oben unter dieser neuen Form: "
        + (", ".join(f"`{o}` ({missing_alt[o]})" for o in sorted(gone)) or "keine"))
    add("")

    # --- 11 CSV-Gegenprobe -----------------------------------------------
    if args.csv_box1:
        add("## 11 CSV-Export Box 1 gegen Blatt Box 1 der NEU-Mappe")
        add("")
        csv = pd.read_csv(args.csv_box1, dtype=str, keep_default_na=False)
        box1 = pd.read_excel(args.neu, sheet_name="Box 1", dtype=object)
        add(f"CSV Zeilen {len(csv)}, XLSX Blatt Box 1 Zeilen {len(box1)}.")
        add("")
        add(f"CSV-Spalten `{list(csv.columns)}`")
        add("")
        n = min(len(csv), len(box1))
        cols = min(csv.shape[1], box1.shape[1])
        diffs = []
        for i in range(n):
            for j in range(cols):
                a = cell(csv.iat[i, j])
                b = cell(box1.iat[i, j])
                if a == b:
                    continue
                # Float-Ganzzahl und Excel-Zeitanteil als gleich werten
                if norm_folio(a) == norm_folio(b):
                    continue
                diffs.append((i + 2, csv.columns[j], a, b))
        add(f"Zellabweichungen in den ersten {n} Zeilen und {cols} Spalten: {len(diffs)}.")
        add("")
        add("| CSV-Zeile | Spalte | CSV | XLSX |")
        add("|---:|---|---|---|")
        for i, (r, c, a, b) in enumerate(diffs):
            if i >= args.sample:
                add(f"| ... | | | insgesamt {len(diffs)} |")
                break
            add(f"| {r} | {c} | `{a[:60]}` | `{b[:60]}` |")
        if not diffs:
            add("| (keine) | | | |")
        add("")

    # --- 12 Zweitfassung gegen ALT ---------------------------------------
    if args.zweitfassung:
        add("## 12 Zweitfassung gegen ALT")
        add("")
        df_z, _ = load_verknuepfungen(args.zweitfassung)
        add(f"Zweitfassung `{args.zweitfassung}`, {len(df_z)} Zeilen "
            f"(ALT {len(df_alt)}).")
        add("")
        za = build_key_map(df_z, row_key)
        aa = build_key_map(df_alt, row_key)
        zn = build_key_map(df_z, row_key_norm)
        an = build_key_map(df_alt, row_key_norm)
        add("| Schluessel | nur Zweitfassung | nur ALT |")
        add("|---|---:|---:|")
        add(f"| strikt | {len([k for k in za if k not in aa])} "
            f"| {len([k for k in aa if k not in za])} |")
        add(f"| normalisiert | {len([k for k in zn if k not in an])} "
            f"| {len([k for k in an if k not in zn])} |")
        add("")
        add("| Signatur | Folio | typ | name | ALT | Zweitfassung |")
        add("|---|---|---|---|---|---|")
        shown = 0
        for k in sorted(set(zn) | set(an)):
            pa = {e["payload"] for e in an.get(k, [])}
            pz = {e["payload"] for e in zn.get(k, [])}
            if pa == pz:
                continue
            shown += 1
            if shown > args.sample:
                continue
            add(f"| {k[0]} | `{k[1]}` | `{k[2]}` | `{k[3][:40]}` | "
                f"{sorted(pa)} | {sorted(pz)} |")
        if shown == 0:
            add("| (keine inhaltliche Abweichung) | | | | | |")
        elif shown > args.sample:
            add(f"| ... | | | | | insgesamt {shown} |")
        add("")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text("\n".join(L), encoding="utf-8")
    print(f"Bericht geschrieben: {args.out} ({len(L)} Zeilen)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
